from flask import Flask, session ,render_template, request,jsonify, url_for, redirect, flash, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column
from sqlalchemy import ForeignKey
from sqlalchemy.orm import relationship
from sqlalchemy import Integer, String, Boolean, DateTime
from flask_login import UserMixin, login_user, LoginManager, current_user, logout_user
import os
from werkzeug.utils import secure_filename
from flask_bootstrap import Bootstrap5
from flask_ckeditor import CKEditor
from flask_login import UserMixin, login_user, LoginManager, login_required, current_user, logout_user
import json
from datetime import datetime
import datetime
import openpyxl


utc_now = datetime.datetime.now()
current_date =f"{utc_now.year}-{utc_now.month}-{utc_now.day}" 
''' 
below  import all the forms required , which is defined in the forms.py
''' 


from forms import (
   
    AddStationForm ,
   
    SensorVisualForm ,
    FEHVisualForm,
   SEHVisualForm,
    MainBridgeVisualForm,
   StumpBridgeVisualForm,
    KaptonGluing,
    HvForm,
    IvForm ,
    SensorGluing,
    NeedleMetrologyForm,
    SkeletonTestForm,
    HybridGluingForm,
    ModuleEncapsulationForm ,
    WireBondingForm,
    NoiseTestForm_Ph2_ACF,
    NoiseTestForm_GIPHT,
    
    ModuleData , SensorForm  ,WireBond,
    VTRxIDForm ,VTRxForm,GroundBalancerIDForm ,  GroundBalancerForm,
    FEHForm, SEHForm, MainBridgeForm, StumpBridgeForm, GlueForm, KaptonTapesForm, OpticalFibreForm, WireBonderForm, OtherConsumablesForm
      ,SensorIdListForm, FEHIdListForm, SEHIdListForm, MainBridgeIdListForm, StumpBridgeIdListForm, GlueBatchIdListForm, KaptonTapeIdListForm, OpticalFibreIdListForm, WireBonderDetailsForm, JigIDForm , OtherConsumablesListForm
)
# make this dictionary for more readability once it works 
add_received_materials_forms = [SensorForm,FEHForm, SEHForm, 
                                MainBridgeForm, StumpBridgeForm, 
                                GlueForm, KaptonTapesForm, OpticalFibreForm,  
                                WireBonderForm,JigIDForm,OtherConsumablesForm,
                                VTRxForm,GroundBalancerForm]
material_receiver_data_dict = [{"sensor_id":[]},{"FEH_id":[]},{"SEH_id":[]},
                               {"main_bridge_id":[]},{"stump_bridge_id":[]},
                               {"glue_batch_id":[],"glue_expiry_date":[]},{"kapton_id":[]},{"optical_fibre_id":[]},
                               {"spool_no":[],"wedge_tool_no":[],"expiry_date":[]},{"jig_id":[]},{"other_id":[]}]
current_material_data = material_receiver_data_dict[0]
Material_receiver_ids_forms = [SensorIdListForm, FEHIdListForm, SEHIdListForm, MainBridgeIdListForm, StumpBridgeIdListForm, 
                               GlueBatchIdListForm, KaptonTapeIdListForm, OpticalFibreIdListForm, WireBonderDetailsForm,  
                               JigIDForm,OtherConsumablesListForm,VTRxIDForm,GroundBalancerIDForm]

''' 
now import all the stups related to the database tike the table and db , all are defined in the database_table.html

'''

#from database_table import User , Station , db
#TO DO: Need to change the import statement to import all the tables from database_table.py
from database_table import (
    db ,
    User,
    Station ,
    MaterialReceiverTable,
    VisualInspectionSensorTable,
    VisualInspectionHybridTable,
    VisualInspectionBridgeTable,
    KaptonGluingTable,
    HvIvFormTable,
    SensorGluingTable,
    NeedleMetrologyTable,
    SkeletonTestTable,
    HybridGluingTable,
    ModuleDataTable,
    WireBondingTable,
    BurNimTable
)


from save_form_database import SaveToDataBase


'''
  create the flask object and follow some rules
'''


app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret-key-goes-here'
# define the Folder path here 

# app.config['UPLOAD_FOLDER'] = "static/uploads"

app.config["UPLOAD_WORKFLOW_FILES"]= "static/WORKFLOW_FILES"
ckeditor = CKEditor(app)
Bootstrap5(app)

db_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'DATABASE')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(db_dir, "nisers.db")}'
db.init_app(app)

# Define save_get_file_url function which take the the form file data , save the data in a given folder and returns path
def save_get_file_url(file):
    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_WORKFLOW_FILES'],filename)
    file.save(file_path)
    return file_path

# Your existing functions
def get_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def generate_filename(workflow_step, material_type, file_extension, file_index=None):
    """
    Generates a filename based on the workflow step, material type, and current timestamp.
    
    Args:
        workflow_step (str): The name of the workflow step.
        material_type (str): The type of material.
        file_extension (str): The file extension.
        file_index (int, optional): The index of the file if multiple files are uploaded.
    
    Returns:
        str: The generated filename.
    """
    step_abbr = {
        "Material Receiver": "MR", "Visual Inspection": "VI", "Kapton Gluing": "KG", "HV form": "HV", "IV form": "IV",
        "Sensor Gluing": "SG", "Needle Metrology": "NM", "Skeleton Test": "ST", "Hybrid Gluing": "HG", "Module Encapsulation": "ME",
        "Wire Bonding": "WB", "NoiseTest_Ph2_ACF": "NTPh2", "NoiseTest_GIPHT": "NTGipht", "Burnin Test": "BT"
    }
    material_abbr = {
        "Sensor": "SEN", "FEH": "FEH", "SEH": "SEH", "MainBridge": "MB", "StumpBridge": "SB", "Glue": "GL", "KaptonTapes": "KT",
        "OpticalFibre": "OF", "WireBonder": "WB", "Other": "OTH", "Module": "MOD"
    }
    timestamp = get_timestamp()
    workflow_abbr = step_abbr.get(workflow_step.strip(), "UNK")
    material_abbr = material_abbr.get(material_type.strip(), "UNK")

    if file_index is not None:
        return f"{workflow_abbr}_{material_abbr}_{timestamp}_{file_index}.{file_extension}"
    else:
        return f"{workflow_abbr}_{material_abbr}_{timestamp}.{file_extension}"

# Define extract_general_form_details function to extract general form details from the request form
# and print them for debugging
def extract_general_form_details(form):
    """
    Extracts general form details from the request form and prints them for debugging.
    
    Args:
        form (ImmutableMultiDict): The request form data.
    
    Returns:
        tuple: A tuple containing the extracted form details.
    """
    details = {
        'received_from': form.get('received_from', None),
        'date': form.get('date', 'Unknown'),
        'temperature': form.get('temperature', 'Unknown'),
        'humidity': form.get('humidity', 'Unknown'),
        'dew_point': form.get('dew_point', 'Unknown')
    }

    # Print general form details for debugging
    if details['received_from']:
        print(f"Received From: {details['received_from']}")
    print(f"Date: {details['date']}")
    print(f"Temperature: {details['temperature']}°C")
    print(f"Humidity: {details['humidity']}%")
    print(f"Dew Point: {details['dew_point']}°C")

    return details['received_from'], details['date'], details['temperature'], details['humidity'], details['dew_point']

def handle_file_uploads(request_files, upload_dir, workflow_name, material_types):
    """
    Handles file uploads, saves them to the specified directory, and prints debug information.
    
    Args:
        request_files (ImmutableMultiDict): The files from the request.
        upload_dir (str): The directory to save the uploaded files.
        workflow_name (str): The name of the workflow step.
        material_types (list): The list of material types.
    
    Returns:
        list: A list of filenames of the uploaded files.
    """
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)  # Ensure directory exists

    file_index = 1
    uploaded_files = []
    multiple_files = len(request_files) > 1

    for file_key in request_files:
        file = request_files[file_key]
        if file.filename:
            print(f"Uploaded file: {file_key} => {file.filename}")
            file.seek(0, os.SEEK_END)
            size = file.tell()
            file.seek(0)
            file_extension = file.filename.split('.')[-1]
            material_type = material_types[0] if material_types else "Unknown"

            # Generate filename
            filename = generate_filename(workflow_name, material_type, file_extension, file_index)
            if multiple_files:
                file_index += 1  

            # Save file
            file_path = os.path.join(upload_dir, filename)
            file.save(file_path)

            uploaded_files.append(filename)
            print(f"Uploaded file saved as: {file_path}")
            print(f"File size: {size} bytes")
        else:
            print(f"Empty file upload: {file_key}")

    return uploaded_files

FORM_MAPPING_VISUAL_INSPECTION = {
   "sensor_visual": SensorVisualForm,
    "FEH_visual": FEHVisualForm,
    "SEH_visual": SEHVisualForm,
    "main_bridge_visual": MainBridgeVisualForm,
    "stump_bridge_visual": StumpBridgeVisualForm,
}

    

#Flask log-in  ,this part is crucial for authenticaton 

login_manager = LoginManager()
login_manager.init_app(app)

@login_manager.user_loader
def load_user(username):
    return db.get_or_404(User ,username)

#upload folder for the images
# Configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# create all the tables 

with app.app_context():
    db.create_all()


# this is the home part for unauthenticated user
@app.route('/')
def home():
    if not current_user.is_authenticated:
        return render_template("index.html")
    else:
        return redirect(url_for('secrets'))

# login part login.html is rendered 
@app.route('/login',methods =["GET","POST"])
def login():
    if request.method =="POST":
        username = request.form.get("username")
        password = request.form.get("password")
        # Query the Users table based on the username
        result = db.session.execute(db.select(User).where(User.username == username))
        user = result.scalar()
        # Check if user exists and is active
        if user:
            if not user.is_active:
                flash("Your account is inactive. Please contact support.")
                return redirect(url_for('login'))

            # Verify password
            if check_password_hash(user.password, password):
                login_user(user)
                return redirect(url_for('secrets'))
            else:
                flash("Incorrect password, please try again.")
        else:
            flash("User does not exist, please check your email.")
    return render_template("login.html")

# once the user is authenticated this is the home page , home.html is rendered 
@app.route('/secrets')
@login_required
def secrets():
    return render_template("home.html")

# when you click on the workflow button in the home page this method is called and workflow.html is rendered 
@app.route('/workflow')
@login_required
def work_flow():
    return render_template("workflow.html")

# when you click on the modules in the homepage this method is called and modules.html is rendered 
@app.route('/modules')
@login_required
def show_module():
    data = db.session.query(HvIvFormTable).all()  # Assuming Station is the model
    columns = HvIvFormTable.__table__.columns
    return render_template("show_table.html", table_data=data, columns=columns)
    return render_template("modules.html")

# when you click on a specific module id this method is executed and module_report.html corresponding to that module will# be rendered and it is under ..
@app.route('/module_report')
@login_required
def module_report():
    return render_template("module_report.html")

@app.route('/wire_bonding', methods=['GET', 'POST'])
@login_required
def wire_bonding():
    if request.method == 'POST':
        ###################Debugging start####################################
        print("\n=== FORM DATA RECEIVED ===")
        received_from, date, temperature, humidity, dew_point = extract_general_form_details(request.form)

        print("\n=== FORM DATA RECEIVED ===")
        for key, value in request.form.items():
            print(f"{key}: {value}")

        # Debugging: Print file upload information
        print("\n=== FILE UPLOADS ===")
        for file_key in request.files:
            file = request.files[file_key]
            if file.filename != '':
                print(f"Uploaded file: {file_key} => {file.filename}")
                file.seek(0, os.SEEK_END)
                size = file.tell()
                file.seek(0)
                print(f"File size: {size} bytes")
            else:
                print(f"Empty file upload: {file_key}")
        ###################Debugging end####################################

        module_id = request.form.get('module_id')
        temperature = request.form.get('temperature')
        dewpoint = request.form.get('dewpoint')
        humidity = request.form.get('humidity')
        comment = request.form.get('comment')
        print(f"\nModule ID: {module_id}, Temperature: {temperature}, Dewpoint: {dewpoint}, Humidity: {humidity}")

        image_path = None
        if 'image' in request.files:
            image = request.files['image']
            if image and allowed_file(image.filename):
                filename = secure_filename(image.filename)
                image_path = os.path.join(UPLOAD_FOLDER, filename)
                image.save(image_path)  
                print(f"Image saved to: {image_path}")

        # Function to process correction factors and forces
        def process_parameters(prefix):
            return {
                'delta_height': request.form.get(f'{prefix}_delta_height'),
                'correction_factor_k1': request.form.get(f'{prefix}_correction_factor_k1'),
                'correction_factor_k2': request.form.get(f'{prefix}_correction_factor_k2'),
                'mean_force_1': request.form.get(f'{prefix}_mean_force_1'),
                'mean_force_2': request.form.get(f'{prefix}_mean_force_2'),
                'rms_value': request.form.get(f'{prefix}_rms_value'),
                'standard_deviation': request.form.get(f'{prefix}_standard_deviation')
            }

        #Top and Bottom Parameters
        top_params = process_parameters('top')
        bottom_params = process_parameters('bottom')
        print("\nTop Parameters:", top_params)
        print("\nBottom Parameters:", bottom_params)

        # Process Dynamic Rows
        def process_table_data(prefix):
            data = []
            i = 1
            while True:
                # Check if at least one field exists for this row
                if not request.form.get(f'{prefix}_raw_pull_force_{i}'):
                    break
                
                row_data = {
                    'raw_pull_force': request.form.get(f'{prefix}_raw_pull_force_{i}'),
                    'distance_between_feet': request.form.get(f'{prefix}_distance_between_feet_{i}'),
                    'type_of_break': request.form.get(f'{prefix}_type_of_break_{i}'),
                    'correction_factor': request.form.get(f'{prefix}_correction_factor_{i}'),
                    'corrected_force': request.form.get(f'{prefix}_corrected_force_{i}'),
                    'comment': request.form.get(f'{prefix}_comment_{i}')
                }
                data.append(row_data)
                print(f"\n{prefix.capitalize()} Row {i}:")
                print(json.dumps(row_data, indent=2))
                i += 1
            return data

        print("\nPROCESSING TOP TABLE DATA:")
        top_data = process_table_data('top')
        print("\nPROCESSING BOTTOM TABLE DATA:")
        bottom_data = process_table_data('bottom')

        # Saving in excel file
        try:
            file_path = 'wire_bonding_data.xlsx'
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Save table data to another sheet if needed
        sheet = workbook.create_sheet(title='Top Data')
        sheet.append(['Raw Pull Force', 'Distance Between Feet', 'Type of Break',
                       'Correction Factor', 'Corrected Force', 'Comment'])
        for row in top_data:
            sheet.append(list(row.values()))

        sheet2 = workbook.create_sheet(title='Bottom Data')
        sheet2.append(['Raw Pull Force', 'Distance Between Feet', 'Type of Break',
                       'Correction Factor', 'Corrected Force', 'Comment'])
        for row in bottom_data:
            sheet2.append(list(row.values()))

        workbook.save(file_path)

        data_dict = {
            'module_id': module_id,
            'temperature': temperature,
            'dewpoint': dewpoint,
            'humidity': humidity,
            'comment': comment,
            'image_path': image_path,
        }


        flash("Wire Bonding data submitted successfully!", "success")
        return redirect(url_for('work_flow'))


    form = WireBondingForm()
    return render_template('wire_bonding.html', form=form)


'''
 this portion is the activated when stations is clicked basically it shows all the stations that are available in the database and gives a options to add more stations 
'''

@app.route('/stations')
@login_required
def stations():
    with app.app_context():
        result = db.session.execute(db.select(Station).order_by(Station.id))
        all_stations = result.scalars().all()
        return render_template("all_stations.html",all_stations = all_stations,user_name = current_user.username)
   
'''
This show_form functions adds the new station to the database and redirect to the the method stations page after successfull completion 

'''
@app.route("/station_form" ,methods=["GET","POST"])
@login_required
def show_form():
    form = AddStationForm()
    if form.validate_on_submit():
        f = form.station_img.data
        station_name = form.station_name.data
        station_location = form.station_location.data
        station_created_at = form.station_created_at.data
        station_remarks = form.station_remarks.data
        #img_path = os.path.join(app.config['UPLOAD_FOLDER'], station_img)
        station_is_active = form.station_is_active.data
        station_iteration_number = form.station_iteration_number.data
        station_operator = form.station_operator.data
        filename = secure_filename(f.filename)
        img_path = os.path.join(app.config['UPLOAD_FOLDER'],filename)
        f.save(os.path.join(img_path))
        new_station = Station(station_name = station_name,
                              station_location = station_location,
                              created_at = station_created_at,
                              remarks = station_remarks,
                              img_path = img_path,
                              is_active = station_is_active,
                              iteration_number = station_iteration_number, 
                              operator = station_operator)
        db.session.add(new_station)
#iteration_number = station_iteration_number,
        db.session.commit()
        return redirect(url_for('stations'))
# this steps make sure the operator filed is auto filled with by the current user name 
    if current_user.is_authenticated:
        form.station_operator.data = current_user.username  # Set station_operator to current user's name
        form.station_operator.render_kw = {'readonly': True}
    return render_template("add_station_form.html",form = form)

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('home'))

# this part meant for downloding the report corresponding to a perticular module
@app.route('/download')
@login_required
def download():
    return send_from_directory("static", path="files/cheat_sheet.pdf")


'''
 here the main thing comes , in the workflow page when you select a process , the num gets it's value which is defined in the workflow.html have a look , that number will tell(see the logic below)  which form to show and after we have to save the data entered in the form to the database once the form is validated 
  Also the visual_inspection.html renders the form
'''
def save_get_file_url(file,file_name):
    file_name = secure_filename(file_name)
    try:
        file_path = os.path.join(app.config['UPLOAD_WORKFLOW_FILES'],file_name)
        file.save(file_path)
        return file_name
    except:
        print("Failed to save the file")
        return None
    
def material_receiver_form_dict(form ,filename):
    materials = {"material_type":None,"common_data":None,"material_data":None}
    materials["common_data"] = {"received_from":form.get('received_from', 'Unknown'),"date":form.get('date', 'Unknown'),
                                "temperature":form.get('temperature', 'Unknown'),"humidity":form.get('humidity', 'Unknown'), "dew_point":form.get('dew_point', 'Unknown'),
                                "material_comment":form.get('material_comment', ''),"img_url":None}
    material_data = {"Sensor":None,"FEH":None,"SEH":None,"Main Bridge":None ,"Stump Bridge":None,"Glue":None,"Kapton Tapes":None,
                     "Optical Fibre":None ,"Wire Bonder":None , "Other":[]}
    for file_key in request.files:
            file = request.files[file_key]
            if file.filename:
                # filename = "filename"
                file_name = save_get_file_url(file,filename)
                materials["common_data"]["img_url"]=file_name
            else:
                print(f"Empty file upload: {file_key}")
    materials["material_type"]=form.getlist('material_type[]')[0]
    if form.getlist('material_type[]')[0]=="Glue":
        materials["material_data"]={"material_ids":form.getlist('material_id[]'),"expiry_dates":form.getlist('expiry_date[]')}

    elif form.getlist('material_type[]')[0]=="Wire Bonder":
        materials["material_data"]={"material_ids":form.getlist('material_id[]'),"expiry_dates":form.getlist('expiry_date[]'),"spool_numbers": form.getlist('spool_number[]') ,
                                    "wedge_tool_numbers":form.getlist('wedge_tool_no[]')}
    else :
        print("else")
        materials["material_data"]={"material_ids":form.getlist('material_id[]')}
    return materials
def get_dict_kapton_gluing_form(form,filename):       
    data_dict = {
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d'),
        "sensor_id": form.sensor_id.data,
        "sensor_type": form.sensor_type.data,
        "cooling_points": form.cooling_points.data,
        "part_A_batch_no": form.part_A_batch_no.data,
        "part_B_batch_no": form.part_B_batch_no.data,
        "jig_no": form.jig_id.data,
        "image": save_get_file_url(form.image.data,filename),
        "comment": form.comment.data
    }
    return data_dict
            
def get_dict_hv_form(form,hv_csv_file_name,other_file_name):
    data_dict = {
"temp": form.temp.data,
"humidity": form.humidity.data,
"dew_point": form.dew_point.data,
"working_date": form.working_date.data.strftime('%Y-%m-%d'),
"sensor_id": form.sensor_id.data,
"cooling_points": form.cooling_points.data,
"hv_csv": save_get_file_url(form.hv_csv.data,hv_csv_file_name),  # CSV file upload
"image": save_get_file_url(form.image.data,other_file_name),  # Optional image upload
"comment": form.comment.data,
}
    return data_dict
def get_dict_iv_form(form,iv_csv_file_name,other_file_name):
    data_dict = {
"temp": form.temp.data,
"humidity": form.humidity.data,
"dew_point": form.dew_point.data,
"working_date": form.working_date.data.strftime('%Y-%m-%d'),
"sensor_id": form.sensor_id.data,
"cooling_points": form.cooling_points.data,
"iv_csv": save_get_file_url(form.iv_csv.data,iv_csv_file_name),  # CSV file upload
"image": save_get_file_url(form.image.data,other_file_name),  # Optional image upload
"comment": form.comment.data,
}
    return data_dict
def get_dict_sensor_gluing_form(form,image_file):
    data_dict = {
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "bare_module_id": form.bare_module_id.data,
        "top_sensor_id": form.top_sensor_id.data,
        "bottom_sensor_id": form.bottom_sensor_id.data,
        "main_bridge_id": form.main_bridge_id.data,
        "stump_bridge_id": form.stump_bridge_id.data,
        "module_spacing": form.module_spacing.data,
        "cooling_points": form.cooling_points.data,
        "jig_id": form.jig_id.data,
        "part_A_batch_no": form.part_A_batch_no.data,
        "part_B_batch_no": form.part_B_batch_no.data,
        "image":save_get_file_url(form.image.data,image_file),
        "comment": form.comment.data,
    }
    return data_dict
def get_dict_needle_metrology(form,csv_excel_file,image_file):
    data_dict ={
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "bare_module_id": form.bare_module_id.data,
        "x_coordinate": form.x_coordinate.data,
        "y_coordinate": form.y_coordinate.data,
        "del_theta": form.del_theta.data,
        "csv_excel": save_get_file_url(form.csv_excel.data,csv_excel_file),  
        "image": save_get_file_url(form.image.data,image_file),  
        "comment": form.comment.data
    }
    return data_dict
def get_skeleton_test_form_data(form,file_name):
    return {
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "skeleton_id": form.skeleton_id.data,
        "FEH_L": form.FEH_L.data,
        "FEH_R": form.FEH_R.data,
        "SEH": form.SEH.data,
        "VTRx": form.VTRx.data,
        "ground_balancer_id": form.ground_balancer_id.data,
        "file": save_get_file_url(form.file.data,file_name), 
        "comment": form.comment.data
    }
def get_hybrid_gluing_form_data(form,image_name):
    return {
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "module_id": form.module_id.data,
        "bare_module_id": form.bare_module_id.data,
        "skeleton_id": form.skeleton_id.data,
        "part_A_batch_no": form.part_A_batch_no.data,
        "part_B_batch_no": form.part_B_batch_no.data,
        "image": save_get_file_url(form.image.data,image_name),  # Optional image upload
        "comment": form.comment.data,
    }
def get_module_encapsulation_form_data(form,image_name):
    return {
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "module_id": form.module_id.data,
        "glue_a": form.glue_a.data,
        "glue_b": form.glue_b.data,
        "glue_preparation_time": form.glue_preparation_time.data,
        "jig": form.jig.data,
        "station": form.station.data,
        "comment": form.comment.data,
        "img":save_get_file_url(form.img.data,image_name) ,  
    }

def get_noise_test_Ph2_ACF_form_data(form,file_dict):
    return {
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "module_id": form.module_id.data,
        "aldrino_file": save_get_file_url(form.upload_folder1.data,file_dict["aldrino_file"]),  # Aldrino File
        "hv_file": save_get_file_url(form.upload_folder2.data,file_dict["hv_file"]),  # HV File
        "lv_file": save_get_file_url(form.upload_folder3.data,file_dict["lv_file"] ),
        "iv_file": save_get_file_url(form.upload_folder4.data,file_dict["iv_file"]),  # IV File
        "root_file": save_get_file_url(form.upload_folder5.data,file_dict["root_file"]),  # ROOT File
        "comment": form.comment.data,
    }
def get_noise_test_GIPHT_form_data(form,file_dict):
    return {
        "temp": form.temp.data,
        "humidity": form.humidity.data,
        "dew_point": form.dew_point.data,
        "working_date": form.working_date.data.strftime('%Y-%m-%d') if form.working_date.data else None,
        "module_id": form.module_id.data,
        "aldrino_file": save_get_file_url(form.upload_folder1.data,file_dict["aldrino_file"]),  # Aldrino File
        "hv_file": save_get_file_url(form.upload_folder2.data,file_dict["hv_file"]),  # HV File
        "lv_file": save_get_file_url(form.upload_folder3.data,file_dict["lv_file"] ),
        "iv_file": save_get_file_url(form.upload_folder4.data,file_dict["iv_file"]),  # IV File
        "root_file": save_get_file_url(form.upload_folder5.data,file_dict["root_file"]),  # ROOT File
        "comment": form.comment.data,
    }

@app.route('/add_data', methods=["GET", "POST"])
def add_data():
    
    workflow_name = request.args.get('workflow_name', 'Workflow')
    
    sensor_ids = db.session.query(VisualInspectionSensorTable.sensor_id).distinct().all()
    bare_module_ids = db.session.query(SensorGluingTable.bare_module_id).distinct().all()
    module_ids = db.session.query(HybridGluingTable.module_id).distinct().all()
    skeleton_ids = db.session.query(SkeletonTestTable.skeleton_id).distinct().all()

    if workflow_name=="Material Receiver" :
        # Log each field for debugging  
        print("\n=== FORM DATA RECEIVED ===")
        received_from, date, temperature, humidity, dew_point = extract_general_form_details(request.form)
        material_types = request.form.getlist('material_type[]')
        material_comment = request.form.get('material_comment', '')
        print(f"{', '.join(set(material_types))} Comment: {material_comment}")

        print("\n=== FILE UPLOADS ===")
        upload_dir = app.config["UPLOAD_WORKFLOW_FILES"]  # Use configured upload path
        uploaded_files = handle_file_uploads(request.files, upload_dir, workflow_name, material_types)
        print("\n=== MATERIAL ENTRIES ===")

        # Retrieve multiple materials
        material_ids = request.form.getlist('material_id[]')
        expiry_dates = request.form.getlist('expiry_date[]')
        spool_numbers = request.form.getlist('spool_number[]')
        wedge_tool_numbers = request.form.getlist('wedge_tool_no[]')

        # Ensure all entries are printed
        
        for i in range(len(material_ids)):  # Iterate over all rows
            material_id = material_ids[i] if i < len(material_ids) else "Unknown"
            material_type = material_types[i] if i < len(material_types) else "Unknown"
            expiry_date = expiry_dates[i] if i < len(expiry_dates) else "N/A"
            spool_no = spool_numbers[i] if i < len(spool_numbers) else "N/A"
            wedge_tool_no = wedge_tool_numbers[i] if i < len(wedge_tool_numbers) else "N/A"

            print(f"\nMaterial Entry {i + 1}:")
            print(f"{material_type} ID: {material_id}")

            if material_type == "Glue":
                print(f"Expiry Date: {expiry_date}")

            elif material_type == "Wire Bonder":
                print(f"Spool Number: {spool_no}, Wedge Tool No.: {wedge_tool_no}, Expiry Date: {expiry_date}")

        return render_template("material_reciever.html")
    
    elif workflow_name == "Visual Inspection":
        return render_template("visual_type.html")
        form = VisualInspection()
        inspection_number = None
        if form.validate_on_submit():
            inspection_number = int(form.inspection_type.data)
            if inspection_number == 1:
                return redirect(url_for('sensor_inspection'))
            elif inspection_number == 3:
                return redirect(url_for('hybrid_inspection'))
            elif inspection_number == 2:
                return redirect(url_for('bridge_inspection'))
        return render_template("visual_inspection.html", form=form, process_name=workflow_name) 
            

        return render_template("visual_type.html")

    elif workflow_name == "Kapton Gluing":
        form = KaptonGluing()
        # add sensor ids from VisualInspectionsencor table to the sensor id choices 
        #sensor_ids = db.session.query(VisualInspectionSensorTable.sensor_id).distinct().all()
        # --------------- replace the below choices from the table -------------------
        sensor_ids = [f"sensor_{i}" for i in range(1, 11)]
        jig_ids = [f"jig_{i}" for i in range(1, 11)]
        glue_ids = [f"glue_{i}" for i in range(1, 11)]
        form.sensor_id.choices = [(sensor_id, sensor_id) for sensor_id in sensor_ids]
        form.part_A_batch_no.choices = [(glue_id, glue_id) for glue_id in glue_ids]
        form.part_B_batch_no.choices = [(glue_id, glue_id) for glue_id in glue_ids]
        form.jig_id.choices = [(jig_id, jig_id) for jig_id in jig_ids]
        if form.validate_on_submit():
            filename = "kapton_glue.png"
            data_dict = get_dict_kapton_gluing_form(form,filename)
            print(data_dict)
            #SaveToDataBase().save_kapton_gluing_form(form,db, app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name == "Hv Form":
        form = HvForm()
        sensor_ids = [f"sensor_{i}" for i in range(1, 11)]
        form.sensor_id.choices = [(sensor_id, sensor_id) for sensor_id in sensor_ids]
        # form.sensor_id.choices = [(sensor.sensor_id, sensor.sensor_id) for sensor in sensor_ids]
        if form.validate_on_submit():
            hv_csv_file_name = "hv_csv.csv"
            other_file_name = "hv_other.csv"
            dict_hv_data = get_dict_hv_form(form,hv_csv_file_name,other_file_name)
            print(dict_hv_data)
            #SaveToDataBase().save_hv_iv_form(form, db, app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name == "Iv Form":
        form = IvForm()
        sensor_ids = [f"sensor_{i}" for i in range(1, 11)]
        form.sensor_id.choices = [(sensor_id, sensor_id) for sensor_id in sensor_ids]
        # form.sensor_id.choices = [(sensor.sensor_id, sensor.sensor_id) for sensor in sensor_ids]
        if form.validate_on_submit():
            iv_csv_file_name = "iv_csv.csv"
            other_file_name = "iv_other.png"
            dict_iv_data = get_dict_iv_form(form,iv_csv_file_name,other_file_name)
            print("iv test data ",dict_iv_data)
            #SaveToDataBase().save_hv_iv_form(form, db, app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name == "Sensor Gluing":
        form = SensorGluing()
        sensor_ids = [f"sensor_{i}" for i in range(1, 11)]
        jig_ids = [f"jig_{i}" for i in range(1, 11)]
        glue_ids = [f"glue_{i}" for i in range(1, 11)]
        main_bridge_ids = [f"main_bridge_{i}" for i in range(1, 11)]  # New list for main_bridge_ids
        stump_bridge_ids = [f"stump_bridge_{i}" for i in range(1, 11)]  # New list for stump_bridge_ids

        # Assigning choices to the form fields
        form.part_A_batch_no.choices = [(glue_id, glue_id) for glue_id in glue_ids]
        form.part_B_batch_no.choices = [(glue_id, glue_id) for glue_id in glue_ids]
        form.jig_id.choices = [(jig_id, jig_id) for jig_id in jig_ids]
        form.top_sensor_id.choices = [(sensor_id, sensor_id) for sensor_id in sensor_ids]
        form.bottom_sensor_id.choices = [(sensor_id, sensor_id) for sensor_id in sensor_ids]

        # Adding choices for main_bridge_id and stump_bridge_id
        form.main_bridge_id.choices = [(main_bridge_id, main_bridge_id) for main_bridge_id in main_bridge_ids]
        form.stump_bridge_id.choices = [(stump_bridge_id, stump_bridge_id) for stump_bridge_id in stump_bridge_ids]
        if form.validate_on_submit():
            image_name = "sensor_gluing_img.png"
            dict_sensor_gluing_data = get_dict_sensor_gluing_form(form,image_name)
            print(dict_sensor_gluing_data)
            #SaveToDataBase().save_sensor_gluing_form(form, db, app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name =="Needle Metrology":
        form = NeedleMetrologyForm()
        bare_module_ids = [f"bare_module_{i}" for i in range(1, 11)]  # List for bare_module_ids

        # Assigning choices to the bare_module_id field
        form.bare_module_id.choices = [(bare_module_id, bare_module_id) for bare_module_id in bare_module_ids]
        # form.bare_module_id.choices = [(bare_module.bare_module_id, bare_module.bare_module_id) for bare_module in bare_module_ids]
        if form.validate_on_submit():
            csv_excel_file = "Needle_csv_excel.png"
            image_file = "image_file.png"
            data_needle_form = get_dict_needle_metrology(form,csv_excel_file,image_file)
            print(data_needle_form)
            #SaveToDataBase().save_needle_metrology_form( form, db, app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name == "Skeleton Test":
        form = SkeletonTestForm()
        feh_l_ids = [f"FEH_L_{i}" for i in range(1, 11)]
        feh_r_ids = [f"FEH_R_{i}" for i in range(1, 11)]
        seh_ids = [f"SEH_{i}" for i in range(1, 11)]
        vtrx_ids = [f"VTRx+_{i}" for i in range(1, 11)]
        ground_balancer_ids = [f"ground_balancer_{i}" for i in range(1, 11)]
        form.FEH_L.choices = [(feh_l_id, feh_l_id) for feh_l_id in feh_l_ids]
        form.FEH_R.choices = [(feh_r_id, feh_r_id) for feh_r_id in feh_r_ids]
        form.SEH.choices = [(seh_id, seh_id) for seh_id in seh_ids]
        form.VTRx.choices = [(vtrx_id, vtrx_id) for vtrx_id in vtrx_ids]
        form.ground_balancer_id.choices = [(ground_balancer_id, ground_balancer_id) for ground_balancer_id in ground_balancer_ids]
        if form.validate_on_submit():
            file_name = "skeleton.png"
            skeleton_dict = get_skeleton_test_form_data(form,file_name)
            print(skeleton_dict)
            #SaveToDataBase().save_skeleton_test_form( form, db, app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name == "Hybrid Gluing":
        form = HybridGluingForm()
        # Example lists of options for each field (can be customized)
        bare_module_ids = [f"bare_module_{i}" for i in range(1, 11)]
        skeleton_ids = [f"skeleton_{i}" for i in range(1, 11)]
        part_A_batch_nos = [f"partA_{i}" for i in range(1, 11)]  # Example values for part_A_batch_no
        part_B_batch_nos = [f"partB_{i}" for i in range(1, 11)]  # Example values for part_B_batch_no

        # Assigning choices to the form fields
        form.bare_module_id.choices = [(bare_module_id, bare_module_id) for bare_module_id in bare_module_ids]
        form.skeleton_id.choices = [(skeleton_id, skeleton_id) for skeleton_id in skeleton_ids]
        form.part_A_batch_no.choices = [(part_A_batch_no, part_A_batch_no) for part_A_batch_no in part_A_batch_nos]
        form.part_B_batch_no.choices = [(part_B_batch_no, part_B_batch_no) for part_B_batch_no in part_B_batch_nos]
        if form.validate_on_submit():
            image_name = "hybrid_gluing.png"
            hybrid_data_dict = get_hybrid_gluing_form_data(form,image_name)
            print(hybrid_data_dict)
            #SaveToDataBase().save_hybrid_gluing_form( form ,db ,app.config['UPLOAD_WORKFLOW_FILES'])
            return redirect(url_for('work_flow'))
    elif workflow_name == "Module Encapsulation":
        form = ModuleEncapsulationForm()
        module_ids = [f"module_{i}" for i in range(1, 11)]
        glue_a_ids = [f"glueA_{i}" for i in range(1, 11)]
        glue_b_ids = [f"glueB_{i}" for i in range(1, 11)]
        jig_ids = [f"jig_{i}" for i in range(1, 11)]
        station_ids = [f"station_{i}" for i in range(1, 11)] 
        form.module_id.choices = [(module_id, module_id) for module_id in module_ids]
        form.glue_a.choices = [(glue_a, glue_a) for glue_a in glue_a_ids]
        form.glue_b.choices = [(glue_b, glue_b) for glue_b in glue_b_ids]
        form.jig.choices = [(jig, jig) for jig in jig_ids]
        form.station.choices = [(station_id, station_id) for station_id in station_ids]
        if form.validate_on_submit():
            image_name = "module_enmcapsualtion.png"
            module_encapsulation_data_dict = get_module_encapsulation_form_data(form,image_name)
            print(module_encapsulation_data_dict)
            return redirect(url_for('work_flow'))

    elif workflow_name == "Wire Bonding": 
        template_name = "wire_bonding.html"  
        return redirect(url_for('wire_bonding'))
    elif workflow_name =="Noise Test Ph2-ACF":
        form = NoiseTestForm_Ph2_ACF()
        module_ids = [f"module_{i}" for i in range(1, 11)]
        form.module_id.choices = [(module_id, module_id) for module_id in module_ids]
        #form.module_id.choices = [(module.module_id, module.module_id) for module in module_ids]
        if form.validate_on_submit():
            file_dict = {
    "aldrino_file": "aldrino.yml",
    "hv_file": "hvfile.png",
    "lv_file": "lvfile.csv",
    "iv_file": "ivfile.txt",
    "root_file": "data.root"
}
            dict_data = get_noise_test_Ph2_ACF_form_data(form,file_dict)
            print(dict_data)
            return redirect(url_for('work_flow'))
    elif workflow_name == "Noise Test GIPHT":
        form = NoiseTestForm_GIPHT()
        # form.module_id.choices = [(module.module_id, module.module_id) for module in module_ids]
        # if form.validate_on_submit():
        module_ids = [f"module_{i}" for i in range(1, 11)]
        form.module_id.choices = [(module_id, module_id) for module_id in module_ids]
        #form.module_id.choices = [(module.module_id, module.module_id) for module in module_ids]
        if form.validate_on_submit():
            file_dict = {
    "aldrino_file": "aldrino.yml",
    "hv_file": "hvfile.png",
    "lv_file": "lvfile.csv",
    "iv_file": "ivfile.txt",
    "root_file": "data.root"
}
            dict_data = get_noise_test_GIPHT_form_data(form,file_dict)
            print(dict_data)
            return redirect(url_for('work_flow'))
    elif workflow_name == "Burnin Test":
        return redirect(url_for('burninTest', num=12))
    
    return render_template("visual_inspection.html", form=form, process_name=workflow_name)

@app.route('/burninTest', methods=["GET", "POST"])
def burninTest():
    idx_num = int(request.args.get("num", 0))
    print(idx_num)
    if idx_num == 12:
        return render_template("BurninTest.html",module_ids =[123,3224,33545,4546])  
    else:
        return "Invalid step number"



def get_dict_form_visual(form,file_name,material_type):
    # dict_materials = {'sensor_visual':"sensor_id",'FEH_visual':"feh_id",'SEH_visual':"seh_id",'main_bridge_visual':"main_bridge_id",
    #                   'stump_bridge_visual':"stump_bridge_id"}
    form_data_dict = {
            "material_type":material_type, 
            "material_id": form.material_id.data,
            "temp": form.temp.data,
            "humidity": form.humidity.data,
            "dew_point": form.dew_point.data,
            "working_date": form.working_date.data.strftime('%Y-%m-%d'),
            "image_url": save_get_file_url(form.image.data,file_name),  # Save file and get the URL
            "comment": form.comment.data
        }
    return form_data_dict
# < ---------------------- Visual Inspection Part ------------------------------
@app.route('/visual_inspection_data', methods=["GET", "POST"])
def visual_inspection_data():
    # get the material type name from  visual_type.html 
    material_type = request.args.get("material_type")
    # get the corresponding form from FORM_MAPPING_VISUAL_INSPECTION (dict)
    material_id_list = ["34rwgey","dgfrtrer3","dfgyuert46","rgf7457"]
    visual_material_form = FORM_MAPPING_VISUAL_INSPECTION[material_type]()
    # get the material primaary key and id as (id,material_id)
    visual_material_form.material_id.choices = [(material_id, material_id) for material_id in material_id_list]
    if visual_material_form.validate_on_submit():
        file_name = "visual.png"
        form_dict_data = get_dict_form_visual(visual_material_form,file_name,material_type)
        print(form_dict_data)
        return redirect(url_for('work_flow'))
    return render_template("visual_inspection.html",form = visual_material_form)
    






if __name__ == "__main__":
    app.run(host='0.0.0.0', port=9999, debug=True)









